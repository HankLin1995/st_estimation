import streamlit as st
import pandas as pd
from tabs import render_falsework_tab, render_channel_tab, render_bridge_tab, render_road_tab, render_wall_tab
import openpyxl
from openpyxl.drawing.image import Image
import os,io
import folium
from streamlit_folium import st_folium
from pyproj import Transformer
import datetime
import requests
import json
from myImage import insert_image
from openpyxl.drawing.image import Image as OpenpyxlImage
from json_test import st_to_json
from datetime import datetime,date

# è‡ªå®šä¹‰ JSON åºåˆ—åŒ–å™¨ï¼Œç”¨äºå¤„ç†æ—¥æœŸå­—æ®µ

def get_basic_price_data():
    unit_price_data = {
        'ææ–™': ['140kg/cm2æ··å‡åœŸ','175kg/cm2æ··å‡åœŸ' ,'210kg/cm2æ··å‡åœŸ', 'é‹¼ç­‹', 'ç”²ç¨®æ¨¡æ¿', 'ä¹™ç¨®æ¨¡æ¿','AC','ç¢çŸ³ç´šé…','CLSM'],
        'å–®åƒ¹': [2700, 2750, 2900, 31, 660, 550,360,790,1460],  
        'å–®ä½': ['m3','m3', 'm3', 'kg', 'm2','m2','m2','m3','m3']
    }
    return pd.DataFrame(unit_price_data)

def get_falsework_price_data():
    unit_price_data = {
        'ææ–™': ['é‹¼æ¿æ¨L=4.5M', 'é‹¼æ¿æ¨L=6M','é‹¼æ¿æ¨L=7M','é‹¼æ¿æ¨L=9M','é‹¼æ¿æ¨L=13M','é‹¼è»Œæ¨L=4M', 'é‹¼è»Œæ¨L=6M','é‹¼è»Œæ¨L=7M','é‹¼è»Œæ¨L=10M','é‹¼è»Œæ¨L=12M'],
        'å–®åƒ¹': [2310,2910,3420,3750,4410,1620,1800,1950,2850,3400],  
        'å–®ä½': ['æ¯é€²è¡Œm']*10
    }
    return pd.DataFrame(unit_price_data)

def get_cost_data(coe_other):

    my_other_cost=0

    cost_data = []
    for key, item in st.session_state['costs'].items():

        my_other_cost=my_other_cost+int(item['total_cost'])
        
        cost_data.append({
            'é …ç›®': item['name'],
            'ç¸½åƒ¹': item['total_cost'],
            'å–®ä½': 'å…ƒ'
        })

    cost_data.append({
        'é …ç›®': 'é›œé …åŠå…¶ä»–',
        'ç¸½åƒ¹': int(my_other_cost*coe_other),
        'å–®ä½': 'å…ƒ' 
    })

    return pd.DataFrame(cost_data)

def generate_cost_report(costs,other_coefficient, indirect_coefficient):
    report = []
    total_direct_cost = 0
    index = 1

    for key, item in costs.items():
        total_cost = item['total_cost']
        if total_cost == 0:
            continue
        
        if key in ['road', 'falsework']:
            description = f"{total_cost:,.0f}å…ƒ"
        else:
            if 'length' in item:
                description = f"{item['unit_cost']:,.0f}å…ƒ/æ¯é€²è¡Œm * {item['length']:,.0f}m"
            elif 'quantity' in item:
                description = f"{item['unit_cost']:,.0f}å…ƒ/æ¯åº§ * {item['quantity']:,.0f}åº§"
            description += f" = {total_cost:,.0f}å…ƒ"
        
        report.append(f"{index}. {item['name']}: {description}")
        total_direct_cost += total_cost
        index += 1

    other_cost=total_direct_cost*other_coefficient
    description = f"{other_cost:,.0f}å…ƒ"
    report.append(f"{index}. é›œé …åŠå…¶ä»–: {description}")

    total_direct_cost=total_direct_cost+other_cost

    indirect_cost = round(total_direct_cost * (1+indirect_coefficient),-3)-total_direct_cost
    total_cost = total_direct_cost + indirect_cost

    report.append(f"\nç›´æ¥å·¥ç¨‹è²» = {total_direct_cost:,.0f}å…ƒ")
    report.append(f"é–“æ¥å·¥ç¨‹è²» = ç›´æ¥å·¥ç¨‹è²» * {indirect_coefficient} = {indirect_cost:,.0f}å…ƒ")
    report.append(f"\nç¸½å·¥ç¨‹è²» = {total_cost:,.0f}å…ƒ")

    st.session_state.totalcost=total_cost
    st.session_state['inf']['job_cost'] = total_cost

    return "\n".join(report)

def check_for_blank_values(data):
    blank_fields = []
    for key, value in data.items():

        if type(value) == str and value=="":
            blank_fields.append(getTitle(key))
        elif type(value) == bool or type(value) == int:
            continue  # ä¸æ£€æŸ¥å¸ƒå°”å€¼æˆ–æ•´æ•°
        elif type(value) == datetime and not value:
            blank_fields.append(getTitle(key))
        elif type(value) == date and not value:
            blank_fields.append(getTitle(key))
        elif type(value) == float and value == 0:
            blank_fields.append(getTitle(key))
        else:
            continue#st.warning(f"æœªå¤„ç†çš„æ•°æ®ç±»å‹: {type(value)} for key: {key}")

    return blank_fields

def getTitle(engname):

    myDict={
        'work_place':'ç¸£å¸‚',
        'work_place2':'é„‰é®å¸‚',
        'work_name':'å·¥ç¨‹åç¨±',
        'work_benefit':'å—ç›Šé¢ç©',
        'work_start_date':'æœ€ä½³æ–½å·¥èµ·å§‹æ—¥æœŸ',
        'work_end_date':'æœ€ä½³æ–½å·¥çµæŸæ—¥æœŸ',
        'work_manage':'åˆ†è™•',
        'work_station':'å·¥ä½œç«™',
        'job_cost':'æ¦‚ä¼°ç¶“è²»',
        'job_length':'æ°´è·¯é•·åº¦'
    }

    return myDict[engname]

def generateXLS(report):

    # Check input empty
    blank_fields = check_for_blank_values(st.session_state['inf'])

# æ˜¾ç¤ºç©ºç™½å€¼
    if blank_fields:
        # st.warning(f"è«‹å°‡ç©ºç™½å…§å®¹å¡«ä¸Š:")
        for r in range(len(blank_fields)):
            st.warning(f"\n\n{blank_fields[r]} :ç©ºç™½!")
   

        # st.warning(f"è«‹å°‡ç©ºç™½å…§å®¹å¡«ä¸Š:\n\n{', '.join(blank_fields)}")
    else:

        st.session_state['inf']['timestamp'] = datetime.now()

        if 'coords' in st.session_state and len(st.session_state['coords']) >= 2:
            # Extract the coordinates
            X1 = st.session_state['coords'][0]['twd97_x']
            Y1 = st.session_state['coords'][0]['twd97_y']
            X2 = st.session_state['coords'][1]['twd97_x']
            Y2 = st.session_state['coords'][1]['twd97_y']
            # X3 = st.session_state['coords'][2]['twd97_x']
            # Y3 = st.session_state['coords'][2]['twd97_y']       
        else:
            # è™•ç† `st.session_state['coords']` å°šæœªåˆå§‹åŒ–æˆ–é•·åº¦ä¸æ˜¯2 çš„æƒ…æ³
            # å¯ä»¥åœ¨é€™è£¡è¨­å®šé©ç•¶çš„é è¨­å€¼æˆ–è€…ç™¼å‡ºè­¦å‘Šè¨Šæ¯
            X1, Y1, X2, Y2 = 0, 0, 0, 0  # è¨­å®šé è¨­å€¼ç‚º0ï¼Œä½ ä¹Ÿå¯ä»¥æ ¹æ“šéœ€æ±‚è¨­å®šå…¶ä»–å€¼
            # st.warning("è«‹æ‰‹å‹•è¼¸å…¥å…©å€‹åº§æ¨™è‡³æ¦‚è¦è¡¨")

        workbook = openpyxl.load_workbook('./template/PLAN.xlsx')
        sheet = workbook["æ¦‚è¦è¡¨"]

        sheet.cell(row=3, column=8).value = report
        sheet.cell(row=16,column=4).value=X1
        sheet.cell(row=17,column=4).value=Y1
        sheet.cell(row=18,column=4).value=X2
        sheet.cell(row=19,column=4).value=Y2
        sheet.cell(row=2,column=1).value=st.session_state['inf']['work_place'] +st.session_state['inf']['work_place2']
        sheet.cell(row=3,column=2).value=st.session_state['inf']['work_station']
        sheet.cell(row=6,column=2).value=st.session_state['inf']['work_name']
        sheet.cell(row=9,column=2).value=st.session_state.totalcost
        sheet.cell(row=14,column=2).value="å—ç›Šé¢ç©"+st.session_state['inf']['work_benefit']+"ha"

        if st.session_state['inf']['work_place_detail']=="å·²å–å¾—ä¸¦ç¢ºèªå¦¥è™•":
            sheet.cell(row=20,column=2).value=' (V)'
        else:
            sheet.cell(row=21,column=2).value=' (V)'

        if st.session_state['inf']['work_water_check']=="æ˜¯":
            sheet.cell(row=22,column=1).value='æ˜¯å¦éœ€é…åˆæ–·æ°´æœŸæ–½å·¥ï¼š     ï¼ˆVï¼‰æ˜¯    ï¼ˆ  ï¼‰å¦'
        else:
            sheet.cell(row=22,column=1).value='æ˜¯å¦éœ€é…åˆæ–·æ°´æœŸæ–½å·¥ï¼š     ï¼ˆ  ï¼‰æ˜¯    ï¼ˆVï¼‰å¦'

        work_start_date = st.session_state['inf']['work_start_date']
        work_end_date = st.session_state['inf']['work_end_date']

        start_date_str = work_start_date.strftime("%Yå¹´%mæœˆ") if work_start_date else "æœªæŒ‡å®š"
        end_date_str = work_end_date.strftime("%Yå¹´%mæœˆ") if work_end_date else "æœªæŒ‡å®š"

        sheet.cell(row=22,column=7).value= f"æœ€ä½³æ–½å·¥æœŸï¼š{start_date_str} ~ {end_date_str}"

        if st.session_state.uploaded_file1 is not None:
            img1_file = io.BytesIO(st.session_state.uploaded_file1.getvalue())
            img1 = OpenpyxlImage(img1_file)
            insert_image(sheet,img1,3,5)
        if st.session_state.uploaded_file2 is not None:
            img2_file = io.BytesIO(st.session_state.uploaded_file2.getvalue())
            img2 = OpenpyxlImage(img2_file)
            insert_image(sheet,img2,14,5)
        if st.session_state.uploaded_file3 is not None:
            img3_file = io.BytesIO(st.session_state.uploaded_file3.getvalue())
            img3 = OpenpyxlImage(img3_file)
            insert_image(sheet,img3,14,8)
        if st.session_state.uploaded_file4 is not None:
            img4_file = io.BytesIO(st.session_state.uploaded_file4.getvalue())
            img4 = OpenpyxlImage(img4_file)
            insert_image(workbook["ä½ç½®åœ–"],img4,3,1)

        sheet = workbook["æå ±æ˜ç´°è¡¨"]

        sheet.cell(row=6,column=1).value=1
        sheet.cell(row=6,column=2).value=st.session_state['inf']['work_name']
        sheet.cell(row=6,column=3).value=st.session_state['costs']['open_channel']['length']
        sheet.cell(row=6,column=5).value=st.session_state['inf']['work_benefit']
        sheet.cell(row=6,column=6).value=st.session_state['inf']['work_place']
        sheet.cell(row=6,column=7).value=st.session_state['inf']['work_place2']
        if st.session_state['inf']['work_place_detail']=="å·²å–å¾—ä¸¦ç¢ºèªå¦¥è™•":
            sheet.cell(row=6,column=8).value='V'
        sheet.cell(row=6,column=9).value=st.session_state.totalcost/1000

        output_file = 'example.xlsx'

        workbook.save(output_file)
        with open(output_file, 'rb') as f:
            bytes_data = f.read()
        btn=st.sidebar.download_button(label='è¨ˆç®—æˆæœä¸‹è¼‰', data=bytes_data, file_name=output_file, type='primary')
        os.remove(output_file)
        savedata()

def savedata():
    json_result = st_to_json(st.session_state)
    # è¨­ç½® Google Apps Script Web æ‡‰ç”¨ç¨‹å¼çš„ URL
    url =st.secrets.GAS_URL 
    with st.sidebar:
        with st.spinner("...è³‡æ–™å„²å­˜ä¸­..."):
            # ç™¼é€ POST è«‹æ±‚ä¸¦å‚³é JSON è³‡æ–™
            response = requests.post(url, data=json_result)
            # æª¢æŸ¥è«‹æ±‚æ˜¯å¦æˆåŠŸ
            if response.status_code == 200:
                st.write("è³‡æ–™å„²å­˜æˆåŠŸ!")
            else:
                st.write("Error:", response.status_code)

def render_page0():

    # st.subheader(":dart:æ­¡è¿ä¾†åˆ°ä¸»é é¢")

    col1,col2,col3=st.columns([5,1,5])

    with col1:

        with open("./md/SP.md", "r", encoding="utf-8") as file:
            markdown_text = file.read()

        st.markdown(markdown_text)


    with col3:

        with open("./md/log.md", "r", encoding="utf-8") as file:
            markdown_text = file.read()

        st.markdown(markdown_text,True) 

    st.session_state.current_page = 'render_page0'

def render_page1():

        # æ£€æŸ¥å¹¶ç¡®ä¿æ—¥æœŸç±»å‹æˆ–ä¸ºç©º
    work_start_date = st.session_state['inf']['work_start_date']
    work_end_date = st.session_state['inf']['work_end_date']

    if isinstance(work_start_date, (str,)):
        work_start_date = None

    if isinstance(work_end_date, (str,)):
        work_end_date = None

    st.subheader("å·¥ç¨‹åŸºæœ¬è³‡æ–™å¡«å ±")

    col1,col2=st.columns([1,1])

    with col1:

        col1_left,col1_right=st.columns([1,1])

        with col1_left:
            st.session_state['inf']['work_place'] = st.text_input("ç¸£å¸‚åˆ¥", value=st.session_state['inf']['work_place'])
            # st.session_state['inf']['work_place2'] = st.text_input("é„‰é®å¸‚åˆ¥", value=st.session_state['inf']['work_place2'])
            st.session_state['inf']['work_manage'] = st.selectbox("åˆ†è™•", options=["æ–—å…­åˆ†è™•","è¥¿èºåˆ†è™•","è™å°¾åˆ†è™•","åŒ—æ¸¯åˆ†è™•","æ—å…§åˆ†è™•"])
        with col1_right:
            st.session_state['inf']['work_place2'] = st.text_input("é„‰é®å¸‚åˆ¥", value=st.session_state['inf']['work_place2'])
            # st.session_state['inf']['work_manage'] = st.selectbox("åˆ†è™•", options=["æ–—å…­åˆ†è™•","è¥¿èºåˆ†è™•","è™å°¾åˆ†è™•","åŒ—æ¸¯åˆ†è™•","æ—å…§åˆ†è™•"])
            st.session_state['inf']['work_station'] = st.text_input("å·¥ä½œç«™", value=st.session_state['inf']['work_station'])
        st.session_state['inf']['work_name'] = st.text_input("æ°´è·¯åç¨±", value=st.session_state['inf']['work_name'])
        st.session_state['inf']['work_benefit'] = st.text_input("å—ç›Šé¢ç©(ha)", value=st.session_state['inf']['work_benefit'])
        
        st.markdown("---")
        st.session_state['inf']['work_place_water'] = st.selectbox("æ°´è·¯ç”¨åœ°", options=["è™•æœ‰åœ°","ç§æœ‰åœ°","å…¬æœ‰åœ°"])
        st.session_state['inf']['work_place_detail'] = st.selectbox("å·¥ç¨‹ç”¨åœ°", options=["å·²å–å¾—ä¸¦ç¢ºèªå¦¥è™•","å°šæœªå–å¾—æˆ–å°šæœªå¦¥è™•"])
        # st.session_state['inf']['work_place_detail'] = st.radio("å·¥ç¨‹ç”¨åœ°", options=["å·²å–å¾—ä¸¦ç¢ºèªå¦¥è™•","å°šæœªå–å¾—æˆ–å°šæœªå¦¥è™•"], 
                                                            # index=0 if st.session_state['inf']['work_place_detail'] != "" else 1)
        # st.session_state['inf']['work_water_check'] = st.selectbox("æ˜¯å¦éœ€è¦é…åˆæ–·æ°´æœŸæ–½å·¥",options=["æ˜¯","å¦"], value=st.session_state['inf']['work_water_check'])
        st.session_state['inf']['work_water_check'] = st.checkbox("éœ€è¦é…åˆæ–·æ°´æœŸæ–½å·¥", value=st.session_state['inf']['work_water_check'])
        st.markdown("---")
        st.session_state['inf']['work_start_date'] = st.date_input("æœ€ä½³æ–½å·¥èµ·å§‹æ—¥æœŸ", value=work_start_date)
        st.session_state['inf']['work_end_date'] = st.date_input("æœ€ä½³æ–½å·¥çµæŸæ—¥æœŸ", value=work_end_date)

    with col2:

        tab1, tab2, tab3, tab4 = st.tabs(["ç¾åœ°è¿‘ç…§", "ç¾åœ°é ç…§", "è¨­è¨ˆç°¡åœ–", "ä½ç½®åœ–"])

        with tab1:
            if  st.session_state.uploaded_file1 is not None:
                uploaded_file1 = st.file_uploader("ç¾åœ°è¿‘ç…§", type=["png", "jpg", "jpeg"], key='upload1')
                if uploaded_file1 is not None:
                    st.session_state.uploaded_file1 = uploaded_file1
                    st.image(st.session_state.uploaded_file1, caption="ç¾åœ°è¿‘ç…§", use_column_width=True)
                else:
                    st.image(st.session_state.uploaded_file1, caption="ç¾åœ°è¿‘ç…§", use_column_width=True)
            else:
                uploaded_file1 = st.file_uploader("ç¾åœ°è¿‘ç…§", type=["png", "jpg", "jpeg"], key='upload1')
                if uploaded_file1 is not None:
                    st.session_state.uploaded_file1 = uploaded_file1
                    st.image(st.session_state.uploaded_file1, caption="ç¾åœ°è¿‘ç…§", use_column_width=True)

        with tab2:
            if  st.session_state.uploaded_file2 is not None:
                uploaded_file2 = st.file_uploader("ç¾åœ°é ç…§", type=["png", "jpg", "jpeg"], key='upload2')
                if uploaded_file2 is not None:
                    st.session_state.uploaded_file2 = uploaded_file2
                    st.image(st.session_state.uploaded_file2, caption="ç¾åœ°é ç…§", use_column_width=True)
                else:
                    st.image(st.session_state.uploaded_file2, caption="ç¾åœ°é ç…§", use_column_width=True)
            else:
                uploaded_file2 = st.file_uploader("ç¾åœ°é ç…§", type=["png", "jpg", "jpeg"], key='upload2')
                if uploaded_file2 is not None:
                    st.session_state.uploaded_file2 = uploaded_file2
                    st.image(st.session_state.uploaded_file2, caption="ç¾åœ°é ç…§", use_column_width=True)

        with tab3:
            if st.session_state.uploaded_file3 is not None:
                uploaded_file3 = st.file_uploader("è¨­è¨ˆç°¡åœ–", type=["png", "jpg", "jpeg"], key='upload3')
                if uploaded_file3 is not None:
                    st.session_state.uploaded_file3 = uploaded_file3
                    st.image(st.session_state.uploaded_file3, caption="è¨­è¨ˆç°¡åœ–", use_column_width=True)
                else:
                    st.image(st.session_state.uploaded_file3, caption="è¨­è¨ˆç°¡åœ–", use_column_width=True)
            else:
                uploaded_file3 = st.file_uploader("è¨­è¨ˆç°¡åœ–", type=["png", "jpg", "jpeg"], key='upload3')
                if uploaded_file3 is not None:
                    st.session_state.uploaded_file3 = uploaded_file3
                    st.image(st.session_state.uploaded_file3, caption="è¨­è¨ˆç°¡åœ–", use_column_width=True)

        with tab4:
            if  st.session_state.uploaded_file4 is not None:
                uploaded_file4 = st.file_uploader("ä½ç½®åœ–", type=["png", "jpg", "jpeg"], key='upload4')
                if uploaded_file4 is not None:
                    st.session_state.uploaded_file4 = uploaded_file4
                    st.image(st.session_state.uploaded_file4, caption="ä½ç½®åœ–", use_column_width=True)
                else:
                    st.image(st.session_state.uploaded_file4, caption="ä½ç½®åœ–", use_column_width=True)
            else:
                uploaded_file4 = st.file_uploader("ä½ç½®åœ–", type=["png", "jpg", "jpeg"], key='upload4')
                if uploaded_file4 is not None:
                    st.session_state.uploaded_file4 = uploaded_file4
                    st.image(st.session_state.uploaded_file4, caption="ä½ç½®åœ–", use_column_width=True)

def render_page2():

    if len(st.session_state['coords']) ==0:
        st.subheader("1.é»é¸æ¸ é“æ–½ä½œèµ·é»")
    elif len(st.session_state['coords']) ==1:
        st.subheader("2.é»é¸æ¸ é“æ–½ä½œçµ‚é»")
    elif len(st.session_state['coords']) == 2:
        st.subheader("3.é»é¸æœ€ä½³æœƒå‹˜åœ°é»")
    else :
        st.warning("**å¦‚æœè¦é‡æ–°é»é¸è«‹å…ˆæ¸…ç©ºæ‰€æœ‰åº§æ¨™**")

    # å®šç¾©åœ°åœ–çš„åˆå§‹ä½ç½®å’Œç¸®æ”¾ç´šåˆ¥
    initial_location = [23.7089, 120.5406]  # é€™è£¡ä½¿ç”¨å°ä¸­çš„ç¶“ç·¯åº¦
    initial_zoom = 10

    # æ·»åŠ é»æ“Šäº‹ä»¶è™•ç†
    def add_marker(folium_map, lat, lon, label):
        folium.Marker(
            location=[lat, lon],
            popup=f"{label}<br>ç¶“åº¦: {lon}<br>ç·¯åº¦: {lat}",
            icon=folium.Icon(icon="info-sign"),
        ).add_to(folium_map)

    # å‰µå»ºä¸€å€‹ Folium åœ°åœ–
    map = folium.Map(location=initial_location, zoom_start=initial_zoom)

    # é¡¯ç¤ºå„²å­˜çš„æ¨™è¨˜
    for i, coord in enumerate(st.session_state['coords']):
        add_marker(map, coord['lat'], coord['lon'], label=f"é» {i + 1}")

    # é¡¯ç¤º Folium åœ°åœ–ä¸¦æ•æ‰é»æ“Šäº‹ä»¶
    map_data = st_folium(map, width=1000, height=500)

    # å¦‚æœæœ‰é»æ“Šäº‹ä»¶ï¼Œç²å–é»æ“Šçš„ä½ç½®
    if map_data and map_data['last_clicked']:
        lat = map_data['last_clicked']['lat']
        lon = map_data['last_clicked']['lng']
        
        # è½‰æ›åæ¨™ç³»çµ±
        transformer = Transformer.from_crs("epsg:4326", "epsg:3826")
        twd97_x, twd97_y = transformer.transform(lat, lon)

        # é¡¯ç¤ºæš«å­˜çš„åæ¨™
        st.write(f"**TWD97 åæ¨™:** X: {twd97_x}, Y: {twd97_y}")

    st.info("å¦‚æœåœ°åœ–æ‰“é–‹æœ‰å›°é›£çš„è©±ï¼Œéœ€è¦å°‡ç¶²é é‡æ–°é–‹å•Ÿï¼Œä»–éœ€è¦å…ˆä¸‹è¼‰æ±è¥¿ä¸‹ä¾†æ‰èƒ½çœ‹çš„åˆ°ã€‚")


    # é¡¯ç¤ºå„²å­˜æŒ‰éˆ•
    if st.sidebar.button('å„²å­˜åº§æ¨™',type='primary') and len(st.session_state['coords']) < 3 :
        st.session_state['coords'].append({'lat': lat, 'lon': lon, 'twd97_x': twd97_x, 'twd97_y': twd97_y})
        st.rerun()

    # é¡¯ç¤ºå„²å­˜çš„åæ¨™
    if len(st.session_state['coords']) == 3:
        # st.sidebar.warning("å·²ç¶“é¸å–äº†å…©å€‹ä½ç½®",icon="âš ï¸")

        if st.sidebar.button('æ¸…ç©ºæ‰€æœ‰åæ¨™'):
            st.session_state['coords'] = []  # æ¸…ç©ºåæ¨™
            st.rerun()  # é‡æ–°é‹è¡Œæ‡‰ç”¨ä»¥æ›´æ–°é é¢

    with st.sidebar:

        for i, coord in enumerate(st.session_state['coords']):
            st.markdown("---")
            st.markdown(f"### é» {i + 1}")
            st.write(f"X: {coord['twd97_x']}")
            st.write(f"Y: {coord['twd97_y']}")

def render_page3():

    col1, col2, col3 = st.columns([8, 1, 4])

    with col3:
        st.markdown("#### :small_blue_diamond: åŸºæœ¬å–®åƒ¹")
        with st.expander("å¸¸è¦‹å¤§å®—ç‰©è³‡"):
            edited_unit_price_df = st.data_editor(get_basic_price_data(), hide_index=True)
        with st.expander("é‹¼ç‰ˆæ¨ã€é‹¼è»Œæ¨"):
            edited_falsework_price_df = st.data_editor(get_falsework_price_data(), hide_index=True)
        
    with col1:
        st.markdown("#### :small_blue_diamond: å·¥ç¨‹é …ç›®")

        tab_names = ["æ¸ é“å·¥ç¨‹", "ç‰ˆæ©‹å·¥ç¨‹", "é“è·¯å·¥ç¨‹", "ç‰ˆæ¨å·¥ç¨‹", "æ“‹åœŸç‰†"]
        tabs = st.tabs(tab_names)
        with tabs[0]:
            render_channel_tab(edited_unit_price_df)
        with tabs[1]:
            render_bridge_tab(edited_unit_price_df)
        with tabs[2]:
            render_road_tab(edited_unit_price_df)
        with tabs[3]:
            render_falsework_tab(edited_falsework_price_df)
        with tabs[4]:
            render_wall_tab(edited_unit_price_df)

    with col3:
        with st.expander(":globe_with_meridians: **ä¼°ç®—æˆæœ**",True):
            coe_other = st.number_input(":star: **é›œé …è²»ç”¨ä¿‚æ•¸**", min_value=0.0, value=0.1, step=0.05)
            cost_df = get_cost_data(coe_other)
            st.dataframe(cost_df, hide_index=True, use_container_width=True)
            coe = st.number_input(":star: **é–“æ¥è²»ç”¨ä¿‚æ•¸**", min_value=0.0, value=0.3, step=0.05)
            sum_cost = cost_df['ç¸½åƒ¹'].sum()
            other_cost = round(sum_cost * (1 + coe), -3) - sum_cost
            total_cost = sum_cost + other_cost
            formatted_sum_cost = f"{sum_cost:,.0f}"
            formatted_other_cost = f"{other_cost:,.0f}"
            formatted_total_cost = f"{total_cost:,.0f}"
            st.write(f"ç›´æ¥è²»ç”¨: {formatted_sum_cost} å…ƒ")
            st.write(f"é–“æ¥è²»ç”¨: {formatted_other_cost} å…ƒ")
        st.markdown(f"##### :large_orange_diamond: ç¸½è²»ç”¨ç‚º {formatted_total_cost} å…ƒ")
        if total_cost != 0:
            with st.sidebar:
                if st.button("å·¥ç¨‹æ¦‚è¦è¡¨", type="primary"):
                    report = generate_cost_report(st.session_state['costs'], coe_other,coe)
                    st.text(report)
                    generateXLS(report)
                    
                    # st.json(st.session_state)

def storeMSG(username, email, txt):

    GAS_URL = st.secrets.GAS_URL_NOTIFY

    data = {
        'username': username,
        'email': email,
        'content': txt
    }

    try:
        response = requests.post(GAS_URL, json=data)
        response.raise_for_status()  # This will raise an HTTPError if the HTTP request returned an unsuccessful status code
        # print(response.json())
        return response.json()
    except requests.exceptions.HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')
        return {'success': False, 'error': str(http_err)}
    except Exception as err:
        print(f'Other error occurred: {err}')
        return {'success': False, 'error': str(err)}
    
def session_initialize():

    if 'costs' not in st.session_state:
        st.session_state['costs'] = {
            'open_channel': {'name':'æ¸ é“å·¥ç¨‹','unit_cost': 0, 'length': 0, 'total_cost': 0},
            'bridge': {'name':'ç‰ˆæ©‹å·¥ç¨‹','unit_cost': 0, 'quantity': 0, 'total_cost': 0},
            'wall': {'name':'æ“‹åœŸç‰†','unit_cost': 0, 'length': 0, 'total_cost': 0},
            'road': {'name':'é“è·¯å·¥ç¨‹','unit_cost': 0, 'quantity': 0, 'total_cost': 0},
            'falsework': {'name':'ç‰ˆæ¨å·¥ç¨‹','unit_cost': 0, 'quantity': 0, 'total_cost': 0}
        }

    if 'totalcost' not in st.session_state:
        st.session_state['totalcost'] = 0

    if 'inf' not in st.session_state:
        st.session_state['inf'] = {
            'timestamp': datetime.now(),
            'work_place': 'é›²æ—ç¸£',
            'work_place2': 'æ–—å…­å¸‚',
            'work_manage':'æ–—å…­åˆ†è™•',
            'work_station': 'æ¢…æ—ç«™',
            'work_name': 'OOå°çµ¦',
            'work_benefit': '20',
            'work_place_water':'è™•æœ‰åœ°',
            'work_place_detail': 'å·²å–å¾—ä¸¦ç¢ºèªå¦¥è™•',
            'work_water_check':True,
            'work_start_date': '',
            'work_end_date': '',
            'job_length':0,
            'job_cost':0

        }

    if 'current_page' not in st.session_state:
        st.session_state['current_page'] = 'page0'

        # ç”¨ä¾†æš«å­˜é»æ“Šçš„åº§æ¨™
    if 'coords' not in st.session_state:
        st.session_state['coords'] = []

        # åˆå§‹åŒ– session_state ä¸­çš„æ–‡ä»¶å±¬æ€§
    if 'uploaded_file1' not in st.session_state:
        st.session_state.uploaded_file1 = None
    if 'uploaded_file2' not in st.session_state:
        st.session_state.uploaded_file2 = None
    if 'uploaded_file3' not in st.session_state:
        st.session_state.uploaded_file3 = None
    if 'uploaded_file4' not in st.session_state:
        st.session_state.uploaded_file4 = None

def main():

    SYSTEM_VERSION="V1.7.3"

    st.set_page_config(
        page_title="å·¥ç¨‹ä¼°ç®—ç³»çµ±"+SYSTEM_VERSION,
        page_icon="ğŸŒ",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    session_initialize()

    with st.sidebar:
        st.title(":globe_with_meridians: å·¥ç¨‹ä¼°ç®—ç³»çµ± "+SYSTEM_VERSION)
        st.write("é€™æ˜¯ç”¨æ–¼æå ±è¨ˆç•«æ™‚çš„ä¼°ç®—å·¥å…·")
        st.info("ä½œè€…:**æ—å®—æ¼¢**")
        with st.expander(":clapper: å½±ç‰‡æ•™å­¸"):
            st.video("./video/demo.mp4")
        with st.expander(":mega: æ„è¦‹å›é¥‹"):

            if 'submitted' not in st.session_state:
                st.session_state.submitted = False

            username = st.text_input(":small_blue_diamond: å§“å")
            email = st.text_input(":small_blue_diamond: é›»å­éƒµä»¶")
            txt = st.text_area(":small_blue_diamond: å…§å®¹")

            if not st.session_state.submitted:
                if st.button("é€å‡º",type='primary'):

                    storeMSG(username, email, txt)
                    st.balloons()
                    st.toast("æ„Ÿè¬ä½ çš„æ„è¦‹å›å¾©!")
                    st.session_state.submitted = True
                    st.rerun()
            else:
                st.write("**:red[æ„Ÿè¬ä½ çš„æ„è¦‹æä¾›! å¦‚è¦ç¹¼çºŒæä¾›è«‹é‡æ–°æ•´ç†]**")
        # st.markdown("---")
        # st.json(st.session_state)
        st.markdown("---")
        st.subheader("é¸æ“‡é é¢")

        if st.button("0.ç³»çµ±æ“ä½œæµç¨‹"):
            st.session_state.current_page = 'page0'  
        if st.button("1.å·¥ç¨‹åŸºæœ¬è³‡æ–™"):
            st.session_state.current_page = 'page1'
        if st.button("2.å·¥ç¨‹æ–½ä½œä½ç½®"):
            st.session_state.current_page = 'page2'
        if st.button("3.å·¥ç¨‹å…§å®¹æ¦‚è¦"):
            st.session_state.current_page = 'page3'
            
        st.markdown("---")
        st.subheader("æ“ä½œæŒ‰éˆ•")

    if st.session_state.current_page == 'page1':
        render_page1()
    elif st.session_state.current_page == 'page2':
        render_page2()
    elif st.session_state.current_page == 'page3':
        render_page3()
    else:
        render_page0()

    # print(st.session_state)

if __name__ == "__main__":
    main()
    # st.sidebar.json(st.session_state)
