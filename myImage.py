# import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker,TwoCellAnchor
import io

def is_merged_cell(ws, col, row):
    for merged_range in ws.merged_cells.ranges:
        if ws.cell(row=row+1, column=col+1).coordinate in merged_range:
            return merged_range
    return None

def insert_image(worksheet,img,target_row,target_col):

    # img= OpenpyxlImage(uploaded_image)

    col1,row1,col2,row2=(is_merged_cell(worksheet, target_col,target_row).bounds)

    marker1=AnchorMarker(col=col1-1, row=row1-1)
    marker2=AnchorMarker(col=col2, row=row2)
    img.anchor = TwoCellAnchor(_from=marker1, to=marker2)
    worksheet.add_image(img)

# Streamlit application title
# st.title("Excel Image Insertion Tool")

# # File upload
# uploaded_excel = st.file_uploader("Upload Excel file", type=["xlsx"])
# uploaded_image = st.file_uploader("Upload image file", type=["png", "jpg", "jpeg"])
# uploaded_image2 = st.file_uploader("Upload image file2", type=["png", "jpg", "jpeg"])

# print(uploaded_excel)
# print(uploaded_image)

# # Input for sheet name and cell location
# sheet_name = st.text_input("Enter sheet name", "概要表")
# cell_location = st.text_input("Enter cell location for image insertion", "E3")

# if uploaded_excel and uploaded_image and uploaded_image2:
#     # Read Excel file
#     workbook = load_workbook(filename=uploaded_excel)
    
#     # Check if sheet exists
#     if sheet_name not in workbook.sheetnames:
#         st.error(f"Sheet '{sheet_name}' does not exist!")
#     else:

#         worksheet = workbook[sheet_name]
#         # Load image
#         # img = OpenpyxlImage(uploaded_image)
#         # img2=OpenpyxlImage(uploaded_image2)

#         insert_image(worksheet,uploaded_image,3,5)
#         insert_image(worksheet,uploaded_image2,14,5)

#         # Save modified Excel file
#         output = io.BytesIO()
#         workbook.save(output)
#         output.seek(0)
        
#         # Provide download link
#         st.download_button(label="Download modified Excel file",
#                            data=output,
#                            file_name="modified_excel.xlsx")

#         st.success("Image successfully inserted into Excel file!")
